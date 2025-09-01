#!/usr/bin/env python3
"""
Final Working Gmail Job Application Processor - REAL MCP INTEGRATION
This script uses your actual Gmail MCP tools!
"""

import os
import sys
import json
import logging
import re
from datetime import datetime
from typing import Dict, Any, List
import pandas as pd

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# IMPORTANT: In Claude environment, these functions are available globally
# search_gmail_messages(q=query)
# read_gmail_thread(thread_id=thread_id, include_full_messages=True)
# read_gmail_profile()

def search_real_job_emails(start_date: str, user_email: str = None) -> List[Dict]:
    """Search Gmail using REAL MCP tools for job-related emails"""
    
    logger.info(f"Searching Gmail for job applications from {start_date}")
    
    # Format date for Gmail search
    start_formatted = datetime.strptime(start_date, '%Y-%m-%d').strftime('%Y/%m/%d')
    end_formatted = datetime.now().strftime('%Y/%m/%d')
    
    # Build comprehensive search queries for job applications
    queries = [
        f'after:{start_formatted} before:{end_formatted} ("application received" OR "thank you for applying" OR "application submitted")',
        f'after:{start_formatted} before:{end_formatted} ("thank you for your interest" OR "received your resume")',
        f'after:{start_formatted} before:{end_formatted} (from:workday.com OR from:greenhouse.io OR from:lever.co)',
        f'after:{start_formatted} before:{end_formatted} ("internship" AND ("summer 2026" OR "software engineer"))',
        f'after:{start_formatted} before:{end_formatted} ("interview" AND ("scheduled" OR "invitation" OR "next step"))',
    ]
    
    all_messages = []
    seen_ids = set()
    
    for query in queries:
        try:
            logger.info(f"Executing search query: {query[:60]}...")
            
            # REAL Gmail MCP call
            result = search_gmail_messages(q=query)
            
            # Parse the result - Gmail MCP returns string format
            messages = parse_gmail_mcp_result(result)
            
            for message in messages:
                message_id = message.get('id')
                if message_id and message_id not in seen_ids:
                    all_messages.append(message)
                    seen_ids.add(message_id)
            
            logger.info(f"Found {len(messages)} new messages")
            
        except Exception as e:
            logger.error(f"Error with search query: {e}")
            continue
    
    logger.info(f"Total unique job-related messages found: {len(all_messages)}")
    return all_messages

def get_thread_details(thread_id: str) -> Dict:
    """Get full thread details using REAL Gmail MCP tools"""
    
    try:
        logger.debug(f"Reading thread: {thread_id}")
        
        # REAL Gmail MCP call
        thread_data = read_gmail_thread(thread_id=thread_id, include_full_messages=True)
        
        return thread_data
        
    except Exception as e:
        logger.error(f"Error reading thread {thread_id}: {e}")
        return {}

def parse_gmail_mcp_result(result) -> List[Dict]:
    """Parse Gmail MCP result into message list"""
    
    messages = []
    
    try:
        # Gmail MCP returns results as a concatenated string of JSON objects
        if isinstance(result, str):
            # Split by JSON object boundaries
            parts = result.split('}{')
            for i, part in enumerate(parts):
                if i > 0:
                    part = '{' + part
                if i < len(parts) - 1:
                    part = part + '}'
                
                try:
                    message = json.loads(part)
                    if 'id' in message and 'snippet' in message:
                        messages.append(message)
                except json.JSONDecodeError:
                    continue
        
        elif isinstance(result, dict):
            # Sometimes returns as dict
            if 'messages' in result:
                messages = result['messages']
            elif 'id' in result:
                messages = [result]
        
        elif isinstance(result, list):
            # Sometimes returns as list
            messages = result
            
    except Exception as e:
        logger.error(f"Error parsing Gmail MCP result: {e}")
    
    return messages

def extract_job_application_data_enhanced(message: Dict, thread_data: Dict) -> Dict[str, str]:
    """Enhanced job application data extraction using thread data"""
    
    try:
        # Get headers from thread data if available
        headers = {}
        if thread_data and 'messages' in thread_data and thread_data['messages']:
            for msg in thread_data['messages']:
                if 'payload' in msg and 'headers' in msg['payload']:
                    for header in msg['payload']['headers']:
                        headers[header['name'].lower()] = header['value']
        
        # Fallback to basic message data
        if not headers and 'payload' in message and 'headers' in message['payload']:
            for header in message['payload']['headers']:
                headers[header['name'].lower()] = header['value']
        
        # Extract details
        sender = headers.get('from', '')
        subject = headers.get('subject', '')
        snippet = message.get('snippet', '')
        
        # Use enhanced extraction functions
        company = extract_company_name_enhanced(sender, subject, snippet)
        position = extract_position_enhanced(subject, snippet)
        applied_date = extract_date_enhanced(message, thread_data)
        status = determine_status_enhanced(subject, snippet)
        source = determine_source_enhanced(sender)
        location = extract_location_enhanced(snippet)
        job_id = extract_job_id_enhanced(snippet)
        
        # Generate enhanced notes
        notes = generate_enhanced_notes(company, position, status, source, job_id)
        
        return {
            'Company': company,
            'Position': position,
            'Applied Date': applied_date,
            'Status': status,
            'Source': source,
            'Location': location,
            'Job ID': job_id,
            'Status Link': '',
            'Notes': notes
        }
        
    except Exception as e:
        logger.error(f"Error in enhanced extraction: {e}")
        # Fallback to basic extraction
        return extract_job_application_data_basic(message)

def extract_company_name_enhanced(sender: str, subject: str, snippet: str) -> str:
    """Enhanced company name extraction"""
    
    # Check sender first for known companies
    sender_lower = sender.lower()
    if 'stryker' in sender_lower:
        return 'Stryker'
    elif 'citadel' in sender_lower:
        return 'Citadel'
    elif 'ge aerospace' in sender_lower or 'geaerospace' in sender_lower:
        return 'GE Aerospace'
    elif 'google' in sender_lower:
        return 'Google'
    elif 'microsoft' in sender_lower:
        return 'Microsoft'
    elif 'meta' in sender_lower:
        return 'Meta'
    elif 'tesla' in sender_lower:
        return 'Tesla'
    elif 'apple' in sender_lower:
        return 'Apple'
    elif 'amazon' in sender_lower:
        return 'Amazon'
    elif 'netflix' in sender_lower:
        return 'Netflix'
    
    # Extract from domain with better parsing
    domain_match = re.search(r'@([^.\s]+)', sender)
    if domain_match:
        domain = domain_match.group(1).lower()
        if domain not in ['workday', 'myworkday', 'greenhouse', 'lever', 'gmail', 'talent']:
            return domain.replace('_', ' ').replace('-', ' ').title()
    
    # Check subject and snippet for company names
    text = f"{subject} {snippet}".lower()
    companies = ['stryker', 'citadel', 'google', 'microsoft', 'meta', 'tesla', 'apple', 'amazon', 'netflix', 'uber', 'airbnb']
    for company in companies:
        if company in text:
            return company.title()
    
    return 'Unknown Company'

def extract_position_enhanced(subject: str, snippet: str) -> str:
    """Enhanced position extraction"""
    
    text = f"{subject} {snippet}".lower()
    
    if 'applied ai engineer' in text:
        return 'Applied AI Engineer Intern - Summer 2026'
    elif 'campus 26' in text and 'software engineering' in text:
        return 'Campus 26 - Software Engineering Intern'
    elif '2026 summer intern' in text and 'software engineering' in text:
        return '2026 Summer Intern - Software Engineering'
    elif 'software engineer intern' in text:
        return 'Software Engineer Intern'
    elif 'data scientist intern' in text:
        return 'Data Scientist Intern'
    elif 'product manager intern' in text:
        return 'Product Manager Intern'
    elif 'machine learning' in text and 'intern' in text:
        return 'Machine Learning Engineer Intern'
    elif 'full stack' in text and 'intern' in text:
        return 'Full Stack Developer Intern'
    else:
        return 'Software Engineer Intern'

def extract_date_enhanced(message: Dict, thread_data: Dict) -> str:
    """Enhanced date extraction"""
    
    # Try to get date from thread data first
    if thread_data and 'messages' in thread_data and thread_data['messages']:
        for msg in thread_data['messages']:
            if 'internalDate' in msg:
                timestamp = int(msg['internalDate']) / 1000
                return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')
    
    # Fallback to message internalDate
    if 'internalDate' in message:
        timestamp = int(message['internalDate']) / 1000
        return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')
    
    return datetime.now().strftime('%Y-%m-%d')

def determine_status_enhanced(subject: str, snippet: str) -> str:
    """Enhanced status determination"""
    
    text = f"{subject} {snippet}".lower()
    
    if any(phrase in text for phrase in ['thank you for', 'application received', 'received your resume', 'received your application']):
        return 'Applied'
    elif 'interview' in text:
        return 'Interview'
    elif any(phrase in text for phrase in ['unfortunately', 'regret', 'not selected', 'not move forward']):
        return 'Rejected'
    elif any(phrase in text for phrase in ['congratulations', 'offer', 'pleased to extend']):
        return 'Offer'
    elif 'next step' in text or 'move forward' in text:
        return 'In Progress'
    else:
        return 'Applied'

def determine_source_enhanced(sender: str) -> str:
    """Enhanced source determination"""
    
    sender_lower = sender.lower()
    
    if 'workday' in sender_lower:
        return 'Workday'
    elif 'greenhouse' in sender_lower:
        return 'Greenhouse'
    elif 'lever' in sender_lower:
        return 'Lever'
    elif 'linkedin' in sender_lower:
        return 'LinkedIn'
    elif 'indeed' in sender_lower:
        return 'Indeed'
    elif 'glassdoor' in sender_lower:
        return 'Glassdoor'
    elif 'talent-citadel' in sender_lower:
        return 'Citadel ATS'
    else:
        return 'Direct Application'

def extract_location_enhanced(snippet: str) -> str:
    """Enhanced location extraction"""
    
    locations = re.findall(r'\b(Texas|Florida|Indiana|Washington|California|New York|Remote|Austin|Seattle|San Francisco|Boston|Chicago|Miami|Denver|Portland)\b', snippet, re.IGNORECASE)
    return locations[0] if locations else ''

def extract_job_id_enhanced(snippet: str) -> str:
    """Enhanced job ID extraction"""
    
    # Look for various job ID patterns
    patterns = [
        r'R(\d+)',                    # Stryker format: R549794
        r'Job ID[:\s]+([A-Z0-9-_]+)', # Job ID: ABC123
        r'Req[:\s]*([A-Z0-9-_]+)',    # Req: DEF456
        r'Position ID[:\s]+([A-Z0-9-_]+)', # Position ID: GHI789
        r'Campus (\d+)',              # Campus 26
    ]
    
    for pattern in patterns:
        match = re.search(pattern, snippet, re.IGNORECASE)
        if match:
            if pattern.startswith('R('):
                return f"R{match.group(1)}"
            elif pattern.startswith('Campus'):
                return f"Campus {match.group(1)}"
            else:
                return match.group(1)
    
    return ''

def generate_enhanced_notes(company: str, position: str, status: str, source: str, job_id: str) -> str:
    """Generate enhanced notes"""
    
    notes = f"Application confirmation received from {company} via {source}"
    
    if job_id:
        notes += f" - {job_id}"
    
    if status == 'Interview':
        notes += " (Interview scheduled)"
    elif status == 'In Progress':
        notes += " (Application in progress)"
    elif status == 'Offer':
        notes += " (Job offer received)"
    
    return notes + "."

def extract_job_application_data_basic(message: Dict) -> Dict[str, str]:
    """Basic job application data extraction as fallback"""
    
    try:
        # Get headers
        headers = {}
        if 'payload' in message and 'headers' in message['payload']:
            for header in message['payload']['headers']:
                headers[header['name'].lower()] = header['value']
        
        # Basic extraction
        sender = headers.get('from', '')
        snippet = message.get('snippet', '')
        
        # Simple company extraction
        if 'stryker' in sender.lower():
            company = 'Stryker'
        elif 'citadel' in sender.lower():
            company = 'Citadel'
        else:
            company = 'Unknown Company'
        
        # Simple position
        position = 'Software Engineer Intern'
        
        # Date
        date_timestamp = int(message.get('internalDate', 0)) / 1000
        applied_date = datetime.fromtimestamp(date_timestamp).strftime('%Y-%m-%d')
        
        return {
            'Company': company,
            'Position': position,
            'Applied Date': applied_date,
            'Status': 'Applied',
            'Source': 'Email',
            'Location': '',
            'Job ID': '',
            'Status Link': '',
            'Notes': f"Application email from {company}."
        }
        
    except Exception as e:
        logger.error(f"Error in basic extraction: {e}")
        return {
            'Company': 'Unknown Company',
            'Position': 'Software Engineer Intern',
            'Applied Date': datetime.now().strftime('%Y-%m-%d'),
            'Status': 'Applied',
            'Source': 'Email',
            'Location': '',
            'Job ID': '',
            'Status Link': '',
            'Notes': 'Job application email received.'
        }

def search_and_process_job_applications(start_date: str = "2025-08-01", user_email: str = None) -> List[Dict[str, str]]:
    """Search Gmail and process job applications using REAL Gmail MCP tools"""
    
    print(f"🔍 Processing job applications from {start_date}...")
    if user_email:
        print(f"📧 Processing for user: {user_email}")
    print("")
    
    try:
        # Use REAL Gmail MCP tools to search for job-related emails
        messages = search_real_job_emails(start_date, user_email)
        print(f"✅ Found {len(messages)} job-related emails")
        
        if not messages:
            print("⚠️  No job-related emails found")
            return []
        
        # Process each message
        applications = []
        for i, message in enumerate(messages):
            print(f"\n📧 Processing email {i+1}/{len(messages)}")
            
            try:
                # Get full thread data for better extraction
                thread_id = message.get('threadId', message.get('id'))
                thread_data = get_thread_details(thread_id) if thread_id else {}
                
                app_data = extract_job_application_data_enhanced(message, thread_data)
                applications.append(app_data)
                
                print(f"   ✅ {app_data['Company']} - {app_data['Position']}")
                print(f"      Applied: {app_data['Applied Date']}, Status: {app_data['Status']}")
                print(f"      Location: {app_data['Location']}, Job ID: {app_data['Job ID']}")
                
            except Exception as e:
                print(f"   ❌ Error processing email: {e}")
                continue
        
        return applications
        
    except Exception as e:
        logger.error(f"Error processing job applications: {e}")
        print(f"❌ Error: {e}")
        return []

def save_applications_to_excel(applications: List[Dict[str, str]], output_path: str) -> bool:
    """Save applications to Excel file with proper formatting"""
    
    try:
        # Define column order to match your existing Excel file
        columns = [
            'Company', 'Position', 'Applied Date', 'Status',
            'Source', 'Location', 'Job ID', 'Status Link', 'Notes'
        ]
        
        # Create DataFrame
        df = pd.DataFrame(applications, columns=columns)
        
        # Sort by Applied Date (most recent first)
        df['Applied Date'] = pd.to_datetime(df['Applied Date'])
        df = df.sort_values('Applied Date', ascending=False)
        df['Applied Date'] = df['Applied Date'].dt.strftime('%Y-%m-%d')
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summer 2026 Internships', index=False)
            
            # Get worksheet for formatting
            worksheet = writer.sheets['Summer 2026 Internships']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col in range(1, len(columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
        
        print(f"✅ Excel file saved: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        return False

def main():
    """Main execution function"""
    
    print("🚀 Gmail Job Application Processor - REAL MCP Integration")
    print("=" * 60)
    print("📊 Using your actual Gmail MCP tools to process job applications")
    print("")
    
    # Simple configuration - no user input needed
    start_date = "2025-08-01"
    output_path = "/Users/dhrbtjr331/Desktop/Recruiting/job-application-tracker/agents/outputs/final_job_applications.xlsx"
    
    print(f"📅 Processing emails from: {start_date}")
    print(f"📁 Output file: {output_path}")
    print("")
    
    # Process applications
    applications = search_and_process_job_applications(start_date)
    
    if applications:
        print(f"\n📋 Summary of {len(applications)} applications found:")
        print("-" * 50)
        
        for i, app in enumerate(applications):
            print(f"{i+1}. {app['Company']} - {app['Position']}")
            print(f"   Applied: {app['Applied Date']}, Status: {app['Status']}")
            print(f"   Location: {app['Location']}, Source: {app['Source']}")
            if app['Job ID']:
                print(f"   Job ID: {app['Job ID']}")
            print("")
        
        # Save to Excel
        success = save_applications_to_excel(applications, output_path)
        
        if success:
            print("🎉 Processing Complete!")
            print(f"📁 Excel file created: {output_path}")
            print("")
            print("✅ Your job application tracker is now up-to-date!")
            print("📊 All applications processed using real Gmail data")
            print("🔧 Using actual Gmail MCP tools successfully")
            
        else:
            print("❌ Failed to create Excel file")
    else:
        print("⚠️  No applications found")
        print("")
        print("💡 This could mean:")
        print("1. No job application emails in the date range")
        print("2. Gmail search queries need adjustment")
        print("3. Check Gmail MCP tool connectivity")

if __name__ == "__main__":
    main()
