# agents/src/agents/excel_writer.py
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
import os
import logging

class ExcelWriterAgent:
    """Agent responsible for writing job application data to Excel files."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define the exact column structure from your example
        self.columns = [
            'Company',
            'Position', 
            'Applied Date',
            'Status',
            'Source',
            'Location',
            'Job ID',
            'Status Link',
            'Notes'
        ]

    def write_to_excel(self, 
                      job_applications: List[Dict[str, Any]], 
                      file_path: str, 
                      overwrite: bool = False) -> bool:
        """Write job applications to Excel file."""
        try:
            # Check if file exists and handle overwrite logic
            if os.path.exists(file_path) and not overwrite:
                self.logger.warning(f"File {file_path} already exists. Use overwrite=True to replace.")
                return False
            
            # Create directory if it doesn't exist
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Convert job applications to DataFrame
            df = self._create_dataframe(job_applications)
            
            # Write to Excel with proper formatting
            self._write_excel_file(df, file_path)
            
            self.logger.info(f"Successfully wrote {len(job_applications)} job applications to {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error writing to Excel: {e}")
            return False

    def append_to_excel(self, 
                       job_applications: List[Dict[str, Any]], 
                       file_path: str) -> bool:
        """Append new job applications to existing Excel file."""
        try:
            existing_data = []
            
            # Read existing data if file exists
            if os.path.exists(file_path):
                existing_df = pd.read_excel(file_path)
                existing_data = existing_df.to_dict('records')
            
            # Combine existing and new data
            all_applications = existing_data + job_applications
            
            # Remove duplicates based on Company + Position + Applied Date
            all_applications = self._remove_duplicates(all_applications)
            
            # Write combined data
            return self.write_to_excel(all_applications, file_path, overwrite=True)
            
        except Exception as e:
            self.logger.error(f"Error appending to Excel: {e}")
            return False

    def _create_dataframe(self, job_applications: List[Dict[str, Any]]) -> pd.DataFrame:
        """Convert job applications to properly formatted DataFrame."""
        # Ensure all required columns exist
        formatted_data = []
        
        for app in job_applications:
            row = {}
            for col in self.columns:
                # Map from JobApplicationData fields to Excel columns
                if col == 'Company':
                    row[col] = app.get('company', '')
                elif col == 'Position':
                    row[col] = app.get('position', '')
                elif col == 'Applied Date':
                    row[col] = app.get('applied_date', '')
                elif col == 'Status':
                    row[col] = app.get('status', '')
                elif col == 'Source':
                    row[col] = app.get('source', '')
                elif col == 'Location':
                    row[col] = app.get('location', '')
                elif col == 'Job ID':
                    row[col] = app.get('job_id', '')
                elif col == 'Status Link':
                    row[col] = app.get('status_link', '')
                elif col == 'Notes':
                    row[col] = app.get('summary', '')  # AI-generated summary
                else:
                    row[col] = ''
            
            formatted_data.append(row)
        
        # Create DataFrame with proper column order
        df = pd.DataFrame(formatted_data, columns=self.columns)
        
        # Sort by Applied Date (most recent first)
        df['Applied Date'] = pd.to_datetime(df['Applied Date'], errors='coerce')
        df = df.sort_values('Applied Date', ascending=False)
        df['Applied Date'] = df['Applied Date'].dt.strftime('%Y-%m-%d')
        
        return df

    def _write_excel_file(self, df: pd.DataFrame, file_path: str):
        """Write DataFrame to Excel with formatting."""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write the main data
            df.to_excel(writer, sheet_name='Summer 2026 Internships', index=False)
            
            # Get the worksheet for formatting
            worksheet = writer.sheets['Summer 2026 Internships']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col in range(1, len(self.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

    def _remove_duplicates(self, applications: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Remove duplicate applications based on key fields."""
        seen = set()
        unique_applications = []
        
        for app in applications:
            # Create a unique key from company, position, and date
            key = (
                app.get('company', '').lower().strip(),
                app.get('position', '').lower().strip(),
                app.get('applied_date', '').strip()
            )
            
            if key not in seen:
                seen.add(key)
                unique_applications.append(app)
        
        return unique_applications

    def get_existing_applications(self, file_path: str) -> List[Dict[str, Any]]:
        """Read existing applications from Excel file."""
        try:
            if not os.path.exists(file_path):
                return []
            
            df = pd.read_excel(file_path)
            return df.to_dict('records')
            
        except Exception as e:
            self.logger.error(f"Error reading existing file: {e}")
            return []
